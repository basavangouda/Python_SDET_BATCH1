import openpyxl
file="D:\\Python_SDET_March-2024-Batch\\Selenium\\Selenium_Methods\\Data_Driven_Testing\\QACircle_Write.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

# Get total number of rows count
rows=sheet.max_row
print(rows)

# Get total number of column count
column=sheet.max_column
print(column)


#Write some data to the file
for r in range(2,rows+1):
    for c in range(1,column+1):
        sheet.cell(r,c).value="Hello"


workbook.save(file)

# Get total number of rows count
rows=sheet.max_row
print(rows)

# Get total number of column count
column=sheet.max_column
print(column)
