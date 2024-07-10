import openpyxl
file="D:\\Python_SDET_March-2024-Batch\\Selenium\\Selenium_Methods\\Data_Driven_Testing\\QACircle_Write_1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

# Get total number of rows count
rows=sheet.max_row
print(rows)

# Get total number of column count
column=sheet.max_column
print(column)

sheet.cell(2,1).value="Python Automation"
sheet.cell(2,2).value="Python"

sheet.cell(3,1).value="Java Automation"
sheet.cell(3,2).value="Java"

sheet.cell(4,1).value="Java Development"
sheet.cell(4,2).value="Java"


workbook.save(file)

# Get total number of rows count
rows=sheet.max_row
print(rows)

# Get total number of column count
column=sheet.max_column
print(column)